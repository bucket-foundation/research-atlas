"""Wellcome Trust connector, via the 360Giving grants-data download.

The Wellcome Trust publishes its **entire grants list** (awarded since
1 October 2000) as a single 360Giving-conformant XLSX download, licensed
CC-BY-4.0. The "Main Grant List" sheet carries one row per grant: applicants,
recipient org (with a 360Giving org identifier), grant programme, title +
description, planned start/end dates, currency + amount awarded, award date,
and the funding org identifier (GB-CHC-210183 = Wellcome).

Canonical download (versioned by month on Wellcome's CMS):
https://cms.wellcome.org/sites/default/files/<YYYY-MM>/Wellcome-grants-awarded-...xlsx
discoverable from the GrantNav publisher page
https://grantnav.threesixtygiving.org/publisher/360G-wellcome .

This is an **offline-after-download** connector: :meth:`fetch` downloads the
XLSX once into ``data/raw/wellcome/`` (resumable), and :meth:`normalize` reads
it with ``openpyxl`` in read-only/streaming mode. No per-row network.

Maps each grant to the canonical graph:
- Funder(Wellcome) -- one fixed funder row (private foundation / charity).
- Grant -- one row. Amount is GBP (or the row's stated currency) -> USD via a
  documented fixed FX. Unknown amounts stay None (money invariant).
- Organization -- the recipient org, with city/region/country; the 360Giving
  ``Recipient Org:Identifier`` is recorded as source_id; ROR resolved via the
  bulk matcher (backfill pass; off by default at connector level).
- Person -- the lead applicant (PI) and named other applicants (co-PI). No
  ORCID is published, so persons fall back to (source, name) keys.
- Field -- the Wellcome "Grant Programme:Title" (Wellcome's own classification).
- Edges: funder->grant, grant->org(recipient), grant->person(pi/co-pi),
  person->org(affiliation).
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import Iterable, Iterator

from atlas.connectors.base import BROWSER_UA, Connector
from atlas.ror_bulk import RorIndex
from atlas.schema import Row, make_id, now_iso

WELLCOME_CROSSREF_ID = "100004440"  # Wellcome Trust, Crossref Funder Registry
WELLCOME_ROR_ID = "https://ror.org/029chgv08"
WELLCOME_PUBLISHER_URL = "https://grantnav.threesixtygiving.org/publisher/360G-wellcome"
WELLCOME_GRANT_URL_TMPL = "https://grantnav.threesixtygiving.org/grant/{}"
WELLCOME_FUNDING_ORG_ID = "GB-CHC-210183"

# Documented fixed FX. Wellcome reports GBP (the Currency column is ~always GBP);
# normalize to USD with a single auditable rate stamped per grant.
GBP_TO_USD = 1.27
FX_AS_OF = "2026-06-01"

# 360Giving country names -> ISO-3166 alpha-2 for ROR country gating.
_COUNTRY_TO_ISO = {
    "united kingdom": "GB", "england": "GB", "scotland": "GB", "wales": "GB",
    "northern ireland": "GB", "united states": "US",
    "united states of america": "US", "switzerland": "CH", "germany": "DE",
    "france": "FR", "netherlands": "NL", "australia": "AU", "india": "IN",
    "kenya": "KE", "south africa": "ZA", "uganda": "UG", "tanzania": "TZ",
    "canada": "CA", "ireland": "IE", "italy": "IT", "spain": "ES",
    "sweden": "SE", "denmark": "DK", "belgium": "BE", "china": "CN",
    "japan": "JP", "thailand": "TH", "malawi": "MW", "ghana": "GH",
    "nigeria": "NG", "ethiopia": "ET", "vietnam": "VN", "brazil": "BR",
    "austria": "AT", "norway": "NO", "finland": "FI", "portugal": "PT",
    "new zealand": "NZ", "singapore": "SG", "zambia": "ZM", "zimbabwe": "ZW",
    "bangladesh": "BD", "pakistan": "PK", "mexico": "MX", "israel": "IL",
    "gambia": "GM", "senegal": "SN", "indonesia": "ID", "nepal": "NP",
}

# Column headers (Main Grant List sheet).
C_FY = "Financial Year"
C_INTERNAL = "Internal ID"
C_SURNAME = "Applicant Surname"
C_LEAD = "Lead Applicant"
C_OTHERS = "Other Applicant(s)"
C_PROGRAMME = "Grant Programme:Title"
C_ORG = "Recipient Org:Name"
C_CITY = "Recipient Org:City"
C_REGION = "Region"
C_ORG_COUNTRY = "Recipient Org:Country"
C_TITLE = "Title"
C_DESC = "Description"
C_START = "Planned Dates:Start Date"
C_END = "Planned Dates:End Date"
C_CURRENCY = "Currency"
C_AMOUNT = "Amount Awarded"
C_AWARD_DATE = "Award Date"
C_IDENTIFIER = "Identifier"
C_ORG_IDENTIFIER = "Recipient Org:Identifier"

SHEET = "Main Grant List"


def _iso_country(country) -> str | None:
    if not country:
        return None
    return _COUNTRY_TO_ISO.get(str(country).strip().lower())


def _to_iso_date(v) -> str | None:
    """openpyxl gives datetime; coerce to ISO date string. None-safe."""
    if v is None or v == "":
        return None
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.date().isoformat() if isinstance(v, _dt.datetime) else v.isoformat()
    s = str(v).strip()
    return s[:10] if len(s) >= 10 else (s or None)


def _parse_amount(v) -> float | None:
    if v in (None, ""):
        return None
    try:
        val = float(str(v).replace(",", "").replace("£", "").strip())
    except (ValueError, TypeError):
        return None
    return val if val > 0 else None


def _split_others(v) -> list[str]:
    """'Other Applicant(s)' is a ';'-separated list (may be None)."""
    if not v:
        return []
    return [p.strip() for p in str(v).split(";") if p.strip()]


def parse_rows(records: Iterable[dict]) -> Iterator[dict]:
    """Pass-through hook so tests can feed dict rows directly (no XLSX)."""
    yield from records


class WellcomeConnector(Connector):
    source = "wellcome"
    user_agent = BROWSER_UA
    delay = 0.5

    FUNDER_ATLAS_ID = make_id("funder", f"crossref:{WELLCOME_CROSSREF_ID}")

    def __init__(self, *args, ror_index: RorIndex | None = None,
                 resolve_ror: bool = False, **kwargs):
        super().__init__(*args, **kwargs)
        self.resolve_ror = resolve_ror
        self._ror_index = ror_index

    # ----- fetch ---------------------------------------------------------- #

    def fetch(self, xlsx_url: str | None = None, **kwargs) -> Iterable[dict]:
        """Download the Wellcome 360Giving XLSX once. Yields {'xlsx_path': ...}.

        The XLSX is binary, so unlike CSV/HTML connectors we cache the raw bytes
        to ``data/raw/wellcome/grants.xlsx`` (not JSON) and yield its path; the
        normalizer streams it with openpyxl. Resumable: a present file is reused.
        """
        if not xlsx_url:
            xlsx_url = self._discover_xlsx_url()
        dest = self.raw_dir / "grants.xlsx"
        if not dest.exists():
            resp = self.http.get(xlsx_url)
            if resp is None:
                return
            dest.write_bytes(resp.content)
        yield {"xlsx_path": str(dest), "url": xlsx_url}

    def _discover_xlsx_url(self) -> str:
        """Scrape the GrantNav publisher page for the current XLSX download URL."""
        cached = self.load_raw("_publisher_page")
        if cached is None:
            resp = self.http.get(WELLCOME_PUBLISHER_URL)
            cached = resp.text if resp is not None else ""
            self.cache_raw("_publisher_page", cached)
        import re
        m = re.search(r'https://cms\.wellcome\.org/[^\s"\'<>]+\.xlsx', cached)
        if m:
            return m.group(0)
        raise RuntimeError(
            "could not discover Wellcome XLSX URL from the GrantNav publisher "
            "page; pass xlsx_url= explicitly"
        )

    # ----- normalize ------------------------------------------------------ #

    def _iter_xlsx(self, path: Path) -> Iterator[dict]:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        for r in rows:
            if r is None:
                continue
            rec = {header[i]: r[i] for i in range(min(len(header), len(r)))}
            if rec.get(C_INTERNAL) or rec.get(C_IDENTIFIER):
                yield rec
        wb.close()

    def normalize(self, raw_pages: Iterable[dict]) -> Iterator[Row]:
        ts = now_iso()
        seen_orgs: set[str] = set()
        seen_persons: set[str] = set()
        seen_fields: set[str] = set()
        funder_emitted = False

        for page in raw_pages:
            if "xlsx_path" in page:
                records = self._iter_xlsx(Path(page["xlsx_path"]))
            elif "rows" in page:  # test hook: pre-parsed dict rows
                records = parse_rows(page["rows"])
            else:
                continue

            for rec in records:
                identifier = (str(rec.get(C_IDENTIFIER) or "").strip()
                              or str(rec.get(C_INTERNAL) or "").strip())
                if not identifier:
                    continue

                if not funder_emitted:
                    yield Row("funder", {
                        "atlas_id": self.FUNDER_ATLAS_ID,
                        "name": "Wellcome Trust",
                        "short_name": "Wellcome",
                        "country_code": "GB",
                        "funder_type": "nonprofit",
                        "ror_id": WELLCOME_ROR_ID,
                        "crossref_funder_id": WELLCOME_CROSSREF_ID,
                        "homepage": "https://wellcome.org",
                        "source": self.source, "source_id": "wellcome",
                        "source_url": WELLCOME_PUBLISHER_URL, "as_of": ts,
                    })
                    funder_emitted = True

                grant_url = WELLCOME_GRANT_URL_TMPL.format(identifier)
                currency = (str(rec.get(C_CURRENCY) or "").strip() or "GBP").upper()
                amount = _parse_amount(rec.get(C_AMOUNT))
                if currency == "GBP":
                    fx = GBP_TO_USD
                elif currency == "USD":
                    fx = 1.0
                else:
                    fx = None  # unknown currency: do not invent an FX
                usd = round(amount * fx, 2) if (amount is not None and fx) else None
                start = _to_iso_date(rec.get(C_START))
                end = _to_iso_date(rec.get(C_END))
                status = "unknown"
                if end:
                    try:
                        status = ("completed"
                                  if _dt.date.fromisoformat(end) < _dt.date.today()
                                  else "active")
                    except ValueError:
                        status = "unknown"
                grant_id = make_id("grant", self.source, identifier)
                # money invariant: only stamp amount/currency/fx when usd resolved
                emit_money = usd is not None
                yield Row("grant", {
                    "atlas_id": grant_id,
                    "title": (str(rec.get(C_TITLE)).strip()
                              if rec.get(C_TITLE) else None),
                    "abstract": (str(rec.get(C_DESC)).strip()
                                 if rec.get(C_DESC) else None),
                    "amount_original": amount if emit_money else None,
                    "currency": currency if emit_money else None,
                    "amount_usd": usd,
                    "fx_rate_to_usd": fx if emit_money else None,
                    "fx_as_of": FX_AS_OF if emit_money else None,
                    "start_date": start,
                    "end_date": end,
                    "status": status,
                    "program": (str(rec.get(C_PROGRAMME)).strip()
                                if rec.get(C_PROGRAMME) else None),
                    "source": self.source, "source_id": identifier,
                    "source_url": grant_url, "as_of": ts,
                })
                yield Row("funder_grant", {
                    "src_id": self.FUNDER_ATLAS_ID, "dst_id": grant_id,
                    "role": "awarder",
                    "source": self.source, "source_id": identifier,
                    "source_url": grant_url, "as_of": ts,
                })

                # --- Organization (recipient) ---
                org_id = None
                org_name = (str(rec.get(C_ORG)).strip()
                            if rec.get(C_ORG) else None)
                if org_name:
                    country = rec.get(C_ORG_COUNTRY)
                    iso = _iso_country(country)
                    org_ident = (str(rec.get(C_ORG_IDENTIFIER)).strip()
                                 if rec.get(C_ORG_IDENTIFIER) else None)
                    ror_id = ror_cc = ror_city = ror_lat = ror_lon = ror_home = None
                    if self.resolve_ror and self._ror_index is not None:
                        m = self._ror_index.match(org_name, iso)
                        if m is not None:
                            r = m.ror
                            ror_id, ror_cc = r.ror_id, r.country_code
                            ror_city, ror_lat, ror_lon = r.city, r.lat, r.lon
                            ror_home = r.homepage
                    org_id = make_id("organization", ror_id) if ror_id \
                        else make_id("organization", self.source, org_name)
                    if org_id not in seen_orgs:
                        seen_orgs.add(org_id)
                        yield Row("organization", {
                            "atlas_id": org_id,
                            "name": org_name,
                            "ror_id": ror_id,
                            "country_code": ror_cc or iso,
                            "city": ror_city or (str(rec.get(C_CITY)).strip()
                                                 if rec.get(C_CITY) else None),
                            "region": (str(rec.get(C_REGION)).strip()
                                       if rec.get(C_REGION) else None),
                            "org_type": "education",
                            "homepage": ror_home,
                            "lat": ror_lat,
                            "lon": ror_lon,
                            "source": self.source,
                            "source_id": ror_id or org_ident or org_name,
                            "source_url": grant_url, "as_of": ts,
                        })
                    yield Row("grant_org", {
                        "src_id": grant_id, "dst_id": org_id, "role": "recipient",
                        "source": self.source, "source_id": identifier,
                        "source_url": grant_url, "as_of": ts,
                    })

                # --- Persons (lead applicant = PI; others = co-PI) ---
                people: list[tuple[str, str]] = []
                lead = (str(rec.get(C_LEAD)).strip() if rec.get(C_LEAD) else None)
                if lead:
                    people.append((lead, "pi"))
                for other in _split_others(rec.get(C_OTHERS)):
                    people.append((other, "co-pi"))
                seen_in_grant: set[str] = set()
                for full, role in people:
                    if not full or full.lower() in seen_in_grant:
                        continue
                    seen_in_grant.add(full.lower())
                    person_id = make_id("person", self.source, full)
                    if person_id not in seen_persons:
                        seen_persons.add(person_id)
                        parts = full.split()
                        yield Row("person", {
                            "atlas_id": person_id,
                            "full_name": full,
                            "first_name": parts[0] if parts else None,
                            "last_name": parts[-1] if len(parts) > 1 else None,
                            "orcid": None,
                            "openalex_author_id": None,
                            "source": self.source, "source_id": full,
                            "source_url": grant_url, "as_of": ts,
                        })
                    yield Row("grant_person", {
                        "src_id": grant_id, "dst_id": person_id, "role": role,
                        "source": self.source, "source_id": identifier,
                        "source_url": grant_url, "as_of": ts,
                    })
                    if org_id:
                        yield Row("person_org", {
                            "src_id": person_id, "dst_id": org_id,
                            "role": "affiliation",
                            "source": self.source, "source_id": identifier,
                            "source_url": grant_url, "as_of": ts,
                        })

                # --- Field (Grant Programme) ---
                prog = (str(rec.get(C_PROGRAMME)).strip()
                        if rec.get(C_PROGRAMME) else None)
                if prog:
                    field_id = make_id("field", self.source, prog)
                    if field_id not in seen_fields:
                        seen_fields.add(field_id)
                        yield Row("field", {
                            "atlas_id": field_id,
                            "name": prog,
                            "openalex_id": None,
                            "level": "field",
                            "parent_atlas_id": None,
                            "source": self.source, "source_id": prog,
                            "source_url": grant_url, "as_of": ts,
                        })
